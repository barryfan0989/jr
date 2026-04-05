import re
import time
import hashlib
from urllib.parse import urljoin, urlparse, parse_qs

import requests
import pandas as pd
from bs4 import BeautifulSoup

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE = "https://kham.com.tw"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}

session = requests.Session()
session.headers.update(HEADERS)


# =========================
# 基本工具
# =========================
def fetch(url: str, sleep_sec: float = 0.4) -> str:
    """抓取網頁 HTML，盡量修正編碼"""
    resp = session.get(url, timeout=25)
    resp.raise_for_status()

    # 若網站沒正確宣告，改用 requests 自動偵測
    if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
        resp.encoding = resp.apparent_encoding

    time.sleep(sleep_sec)
    return resp.text


def clean_text(text: str) -> str:
    """整理空白與奇怪字元"""
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def make_pk(prefix: str, raw: str) -> str:
    """用內容產生穩定主鍵"""
    digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:8]
    return f"{prefix}{digest}"


def normalize_url(href: str) -> str:
    return urljoin(BASE, href)


def extract_product_id(url: str) -> str:
    try:
        qs = parse_qs(urlparse(url).query)
        return qs.get("PRODUCT_ID", [""])[0]
    except Exception:
        return ""


def normalize_artist_name(name: str) -> str:
    name = clean_text(name).lower()
    return re.sub(r"\s+", " ", name)
def clean_artist_display_name(name: str) -> str:
    """
    清理藝人名稱，只保留前面的藝人名
    例如：
    辛曉琪2026《好好愛 此刻》 -> 辛曉琪
    李翊君《翊起聽見愛》 -> 李翊君
    施孝榮《不是你想的那樣》讚聲 -> 施孝榮
    薛恩Sean《RE:BORN: LEVEL UP》讚聲 -> 薛恩Sean
    """
    name = clean_text(name)

    # 先砍掉書名號/括號後面的內容
    name = re.split(r"[《〈（(【\[]", name)[0]
    name = clean_text(name)

    # 去掉結尾年份，例如：辛曉琪2026 -> 辛曉琪
    name = re.sub(r"(19|20)\d{2}$", "", name).strip()

    return name if name else "未知藝人"


# =========================
# 抓首頁所有活動連結
# =========================
def get_event_links_from_home() -> list[str]:
    """
    只抓新版 application 頁面，避免混入其他結構
    """
    html = fetch(BASE)
    soup = BeautifulSoup(html, "lxml")

    links = set()

    for a in soup.select("a[href]"):
        href = a.get("href", "").strip()
        if not href:
            continue

        full_url = normalize_url(href)

        if (
            "PRODUCT_ID=" in full_url
            and "/application/UTK02/UTK0201_.aspx" in full_url
        ):
            links.add(full_url)

    return sorted(links)


# =========================
# 標題 / 藝人 / 類型判斷
# =========================
BAD_TITLE_EXACT = {
    "寬宏售票系統",
    "寬宏藝術",
    "購票系統",
    "立即購票",
    "我要購票",
    "會員登入",
    "商品專區",
}

BAD_TITLE_PARTS = [
    "票價",
    "票種",
    "注意事項",
    "購票須知",
    "退票方式",
    "退票須知",
    "購票方式",
    "如何確認訂票",
    "網路/手機APP購票",
    "節目介紹",
    "商品介紹",
    "下載傳真訂購單",
    "活動相關異動訊息",
    "主辦單位保留",
]

TITLE_KEYWORDS = [
    "演唱會", "音樂會", "見面會", "粉絲見面會", "舞台劇", "音樂劇",
    "展覽", "特展", "演出", "Concert", "LIVE", "Live", "SHOW", "Show",
    "Package", "VIP Package", "商品"
]


def looks_like_bad_title(text: str) -> bool:
    if not text:
        return True

    if text in BAD_TITLE_EXACT:
        return True

    if any(part in text for part in BAD_TITLE_PARTS):
        return True

    # 太短通常不是標題
    if len(text) < 4:
        return True

    # 太長大多是說明段落
    if len(text) > 80:
        return True

    return False


def score_title_candidate(text: str) -> int:
    score = 0

    if not looks_like_bad_title(text):
        score += 20

    for kw in TITLE_KEYWORDS:
        if kw.lower() in text.lower():
            score += 15

    # 有書名號很像正式標題
    if "《" in text and "》" in text:
        score += 20

    # 有年份常見於活動名
    if re.search(r"\b20\d{2}\b", text):
        score += 10

    # 長度適中加分
    if 8 <= len(text) <= 40:
        score += 15

    # 不要把純系統字樣抓進來
    if "寬宏售票系統" in text:
        score -= 50

    return score


def collect_title_candidates(soup: BeautifulSoup) -> list[str]:
    candidates = []

    # 1. meta / og 標題
    meta_props = [
        ("meta", {"property": "og:title"}),
        ("meta", {"name": "title"}),
        ("meta", {"property": "twitter:title"}),
    ]
    for tag_name, attrs in meta_props:
        node = soup.find(tag_name, attrs=attrs)
        if node:
            content = clean_text(node.get("content", ""))
            if content:
                candidates.append(content)

    # 2. 常見可視標題區
    for selector in [
        "h1", "h2", "h3", "title", "strong", "b",
        "font", "td", "div", "span"
    ]:
        for node in soup.select(selector):
            text = clean_text(node.get_text(" ", strip=True))
            if text:
                candidates.append(text)

    # 去重保序
    seen = set()
    uniq = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            uniq.append(c)

    return uniq


def extract_title(soup: BeautifulSoup, product_id: str) -> str:
    """
    從多個候選中找最像真正活動名稱的文字
    """
    candidates = collect_title_candidates(soup)

    best_text = ""
    best_score = -999

    for text in candidates:
        score = score_title_candidate(text)

        # 若出現「商品-」也保留，因為有些 PRODUCT_ID 就是商品頁
        if text.startswith("商品-"):
            score += 20

        # 有 product/vip package 關鍵字也可能是真標題
        if "VIP Package" in text or "Package" in text:
            score += 10

        if score > best_score:
            best_score = score
            best_text = text

    # 最後清理尾巴
    title = clean_text(best_text)
    title = re.sub(r"\s*[-｜|]\s*寬宏.*$", "", title)
    title = re.sub(r"\s*-\s*KHAM.*$", "", title, flags=re.I)
    title = clean_text(title)

    if not title or looks_like_bad_title(title):
        return f"未提供_{product_id}" if product_id else "未提供"

    return title


def infer_event_category_from_title(title: str) -> str:
    title_lower = title.lower()
    if "商品" in title or "商品-" in title_lower:
        return "商品"
    if "vip package" in title_lower or "package" in title_lower:
        return "VIP套票"
    if "展" in title or "特展" in title:
        return "展覽"
    if "音樂劇" in title:
        return "音樂劇"
    if "演唱會" in title:
        return "演唱會"
    if "音樂會" in title:
        return "音樂會"
    if "見面會" in title:
        return "見面會"
    return "其他"


def extract_artist_from_title(title: str) -> str:
    """
    從標題推藝人。對展覽 / 商品 / VIP 套票比較保守。
    """
    title = clean_text(title)
    if not title or title.startswith("未提供_"):
        return "未知藝人"

    category = infer_event_category_from_title(title)

    # 商品頁通常沒有藝人
    if category == "商品":
        return "未知藝人"

    patterns = [
        r"^\d{4}\s*(.+?)\s*(?:演唱會|音樂會|見面會|粉絲見面會)",
        r"^(.+?)\s*(?:演唱會|音樂會|見面會|粉絲見面會)",
        r"^(.+?)\s+LIVE\b",
        r"^全本音樂劇《(.+?)》",
        r"^音樂劇《(.+?)》",
        r"^《(.+?)》",
        r"^【VIP Package】(.+?)(?:演唱會|音樂會|見面會|VIP Package|Package)",
        r"^(.+?)\s*-\s*展場限定",
        r"^商品-(.+)",
    ]

    for pattern in patterns:
        m = re.search(pattern, title, flags=re.I)
        if m:
            artist = clean_text(m.group(1))
            if artist and len(artist) <= 40:
                # 去掉容易殘留的前後字
                artist = re.sub(r"^[【\[]VIP Package[】\]]", "", artist, flags=re.I).strip()
                artist = re.sub(r"\s*VIP Package$", "", artist, flags=re.I).strip()
                return clean_artist_display_name(artist)

    # 展覽很多沒有藝人，用活動名本身較合理
    if category in {"展覽", "音樂劇"}:
        return clean_artist_display_name(title) if len(title) <= 40 else "未知藝人"

    return "未知藝人"


# =========================
# 票價 / 票種 / 售票時間 / 地點
# =========================
def extract_sale_time(text: str) -> str:
    patterns = [
        r"(全面啟售[:：]?\s*[^。；\n]+)",
        r"(開賣日期[:：]?\s*[^。；\n]+)",
        r"(售票日期[:：]?\s*[^。；\n]+)",
        r"(售票時間[:：]?\s*[^。；\n]+)",
        r"(啟售時間[:：]?\s*[^。；\n]+)",
        r"(購票時間[:：]?\s*[^。；\n]+)",
        r"(開賣時間[:：]?\s*[^。；\n]+)",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return clean_text(m.group(1))
    return ""


def extract_ticket_price(text: str) -> str:
    prices = set()

    for m in re.findall(r"NT\$ ?\d{1,5}", text, flags=re.I):
        prices.add(m.upper().replace(" ", ""))

    for m in re.findall(r"\d{1,5}\s*元", text):
        prices.add(clean_text(m))

    def sort_key(x: str) -> int:
        nums = re.findall(r"\d+", x)
        return int(nums[0]) if nums else 999999

    return "、".join(sorted(prices, key=sort_key))


def extract_ticket_type(text: str) -> str:
    keywords = [
        "全票", "早鳥票", "學生票", "身障票", "套票",
        "VIP", "預售票", "搖滾區", "看台區", "站票",
        "單人票", "雙人票", "青春票"
    ]
    found = []
    for kw in keywords:
        if kw in text:
            found.append(kw)
    return "、".join(dict.fromkeys(found))


def extract_location(text: str) -> str:
    known_patterns = [
        r"(台北小巨蛋)",
        r"(高雄巨蛋)",
        r"(臺北流行音樂中心)",
        r"(台北流行音樂中心)",
        r"(高雄流行音樂中心)",
        r"(台中國家歌劇院)",
        r"(國家戲劇院)",
        r"(國家音樂廳)",
        r"(Legacy Taipei)",
        r"(Zepp New Taipei)",
        r"(Zepp Kaohsiung)",
        r"(台北國際會議中心)",
        r"(台北和平籃球館)",
    ]
    for p in known_patterns:
        m = re.search(p, text)
        if m:
            return m.group(1)

    fallback_patterns = [
        r"(?:演出地點|活動地點|場地|地點)[:：]?\s*([^。；\n]{2,40})",
        r"(?:演出場地|展覽地點|展覽場地)[:：]?\s*([^。；\n]{2,40})",
    ]
    for p in fallback_patterns:
        m = re.search(p, text)
        if m:
            return clean_text(m.group(1))

    return "未提供"


# =========================
# 單頁解析
# =========================
def parse_event_page(url: str) -> dict:
    html = fetch(url)
    soup = BeautifulSoup(html, "lxml")
    page_text = clean_text(soup.get_text(" ", strip=True))

    product_id = extract_product_id(url)
    title = extract_title(soup, product_id)
    artist = extract_artist_from_title(title)
    sale_time = extract_sale_time(page_text)
    ticket_price = extract_ticket_price(page_text)
    ticket_type = extract_ticket_type(page_text)
    location = extract_location(page_text)

    return {
        "活動編號PK": product_id if product_id else make_pk("ACT", url),
        "活動名稱": title,
        "票價": ticket_price,
        "票種": ticket_type,
        "活動時間": "",
        "售票時間": sale_time,
        "活動連結": url,
        "場地名稱": location,
        "藝人名稱": artist,
    }


# =========================
# 建表
# =========================
def build_tables(event_links: list[str]):
    events_raw = []

    for i, link in enumerate(event_links, start=1):
        try:
            print(f"[{i}/{len(event_links)}] 抓取: {link}")
            events_raw.append(parse_event_page(link))
        except Exception as e:
            print(f"抓取失敗: {link} -> {e}")

    # 場地表
    location_map = {}
    locations = []

    # 藝人表（去重）
    artist_map = {}
    artists = []

    # 活動表
    events = []

    # 售票平台表
    platforms = []

    for e in events_raw:
        # 場地去重
        location_name = e["場地名稱"] if e["場地名稱"] else "未提供"
        if location_name not in location_map:
            loc_id = make_pk("LOC", location_name)
            location_map[location_name] = loc_id
            locations.append({
                "場地編號PK": loc_id,
                "場地名稱": location_name,
                "地址": ""
            })

        # 藝人去重
        artist_name = e["藝人名稱"] if e["藝人名稱"] else "未知藝人"
        artist_norm = normalize_artist_name(artist_name)

        if artist_norm not in artist_map:
            artist_id = make_pk("ART", artist_norm)
            artist_map[artist_norm] = artist_id
            artists.append({
                "藝人編號PK": artist_id,
                "藝人名稱": artist_name
            })

        # 活動表
        events.append({
            "活動編號PK": e["活動編號PK"],
            "場地編號FK": location_map[location_name],
            "藝人編號FK": artist_map[artist_norm],
            "活動名稱": e["活動名稱"],
            "票價": e["票價"],
            "票種": e["票種"],
            "活動時間": e["活動時間"],
            "售票時間": e["售票時間"],
            "活動連結": e["活動連結"],
        })

        # 售票平台表
        platforms.append({
            "平台編號PK": make_pk("PLT", f"{e['活動編號PK']}_寬宏售票"),
            "活動編號FK": e["活動編號PK"],
            "藝人編號FK": artist_map[artist_norm],
            "平台名稱": "寬宏售票"
        })

    # 售票平台去重
    unique_platforms = []
    seen_platform = set()
    for p in platforms:
        key = (p["活動編號FK"], p["藝人編號FK"], p["平台名稱"])
        if key not in seen_platform:
            seen_platform.add(key)
            unique_platforms.append(p)

    # 使用者表：公開網站沒有這些資料
    users = [{
        "使用者編號PK": "",
        "ID": "",
        "密碼": ""
    }]

    return users, events, locations, artists, unique_platforms


# =========================
# 輸出 CSV / Excel
# =========================
def save_csv(filename: str, rows: list[dict]):
    df = pd.DataFrame(rows)
    df.to_csv(filename, index=False, encoding="utf-8-sig")
    print(f"已輸出 {filename}")


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24

    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def save_excel(filename: str, users, events, locations, artists, platforms):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        pd.DataFrame(users).to_excel(writer, sheet_name="使用者", index=False)
        pd.DataFrame(events).to_excel(writer, sheet_name="活動", index=False)
        pd.DataFrame(locations).to_excel(writer, sheet_name="活動地點", index=False)
        pd.DataFrame(artists).to_excel(writer, sheet_name="藝人", index=False)
        pd.DataFrame(platforms).to_excel(writer, sheet_name="售票平台", index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            style_worksheet(ws)

    print(f"已輸出 {filename}")


# =========================
# 主程式
# =========================
def main():
    event_links = get_event_links_from_home()
    print(f"共找到 {len(event_links)} 個活動連結")

    users, events, locations, artists, platforms = build_tables(event_links)

    save_csv("使用者.csv", users)
    save_csv("活動.csv", events)
    save_csv("活動地點.csv", locations)
    save_csv("藝人.csv", artists)
    save_csv("售票平台.csv", platforms)

    save_excel("kham_data.xlsx", users, events, locations, artists, platforms)

    print("全部完成！")


if __name__ == "__main__":
    main()